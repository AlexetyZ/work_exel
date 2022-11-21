import os
import openpyxl
import shutil
import re
from slovar import CHVR





class Exel_work:
    def __init__(self):
        pass

    """
    Для сортировки файлов по папкам, на основании общего списка (lib) формата xlsx.
    
    требования для lib:
        -столбец А должен содержать начало имени файлов, а столбец С должен содержать название папки, куда надо файл расфасовать
        -без заголовков, в А1 должно быть 1 значение
    
    PS: не забудь в строчке 37 изменить значение конца файла
    """

    def sort_from_lib(self, path_files, path_file_lib):
        self.path_file_lib = path_file_lib #'/Users/aleksejzajcev/Desktop/Список_планов_КНМ.xlsx'
        self.path_files = path_files  # '/Users/aleksejzajcev/Desktop/выгруженные планы на 2023 без Свердловской области 2023051239'
        self.open_lib()

    """
    Функция для слияния файлов xlsx в папке
    """
    def merge_files_in_directory(self, dir, start_from_row):

        if dir.split('/')[-1] == '.DS_Store':
            return 0
        self.dir = dir

        self.start_from_row = start_from_row
        for file in os.listdir(self.dir):
            print(file)
            # self.delete_last_empty_rows(f'{self.dir}/{file}', start_from_row=self.start_from_row)

        while True:
            files_list = os.listdir(self.dir)

            try:
                for file in files_list:
                    if re.search(r'\._', file):
                        files_list.remove(file)

            except:
                pass

            if len(files_list) > 1:
                print('объединяем')
                self.merge_files(files_list)

            else:
                print('остался 1 файл')
                break


    def delete_last_empty_rows(self, path, start_from_row):
        """
        если нужно удалить лишние пустые строчки в конце файла,
        пользуйся этим, но помни, нужно иметь показательную ячейку,
        пустота в которой будет означать, что вся строка пустая.
        Например, план проверок, если в нем пустая строка 4, то это 100% значит, что вся строка пустая.
        """
        wb = openpyxl.load_workbook(path)
        sh = wb.worksheets[0]
        sh_max_row = sh.max_row
        for row in range(start_from_row, sh_max_row):
            # print(f'{row=} {sh.cell(row=row, column=4).value=}')
            if sh.cell(row=row, column=4).value == None:
                amount = sh_max_row-row+1
                sh.delete_rows(idx=row, amount=amount)
                break
        #
        # print(f'{sh.max_row=}')
        # print(f'{len(sh["D"])=}')
        wb.save(path)






    def coint_knm(self, dir, start_from_row):
        '''
        посчитать количество детский садов и школ по наличию в файле столбца B.
        файл обязательно должен иметь расширение xlsx!
        не забудь закрыть все файлы exel, чтобы избежать нежданчика
        ВНИМАНИЕ, в папке должен быть только 1 файл!!!! сначала сделай merge_files_in_directory
        '''

        file_name = str(dir.split('/')[-1])

        count = 0

        if file_name == '.DS_Store':
            return 0
        else:

            file = os.listdir(dir)[0]

            wb = openpyxl.load_workbook(f'{dir}/{file}')
            sh = wb.worksheets[0]
            det_sad = 0
            shkola = 0

            word_shkola = ["ШКОЛА", "ГИМНАЗИЯ", "ЛИЦЕИ", "ОБЩЕОБРАЗОВАТЕЛЬНОЕ", "ОБЩЕЕ ОБРАЗОВАНИЕ"]
            word_det_sad = ["ДЕТСКИЙ САД", "ДОШКОЛЬНОЕ ОБРАЗОВАТЕЛЬНОЕ"]

            for n, ul in enumerate(sh['B']):
                if n < start_from_row:
                    continue
                ul_name = ul.value
                kind_deyat = sh.cell(row=n + 1, column=6).value
                if ul_name is not None and ul_name != 1:
                    count += 1
                    for i in word_shkola:
                        if re.findall(i, ul_name) or re.findall(i, str(kind_deyat).upper()):
                            shkola += 1
                            break

                    for i in word_det_sad:
                        if re.findall(i, ul_name) or re.findall(i, str(kind_deyat).upper()):
                            det_sad += 1
                            break

            print(file_name, count, shkola, det_sad)

    def count_risk(self, dir, start_from_row):
        """
        Определить сколько проверок у субъекта (ТУ) по классам опасности.
        Рекомендуется использовать для выявления ошибок в указании класса опасности.
        формирует табличку в папке ТУ с названием "ошибки", где указаны количество проверок по классам опасности.
        А так же в папке выгрузки формируется общая табличка с указанием ТУ и количестве проверок по классам опасности
        """

        subject_name = dir.split('/')[-1]
        if subject_name == '.DS_Store':
            return 0


        RISK = {
            'чрезвычайно высокий риск': 0,
            'высокий риск': 0,
            'значительный риск': 0,
            'средний риск': 0,
            'умеренный риск': 0,
            'низкий риск': 0,
        }

        file_0 = os.listdir(dir)

        try:
            file_0.remove('.DS_Store')
        except:
            pass

        file = file_0[0]

        wb = openpyxl.load_workbook(f'{dir}/{file}')
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=start_from_row, max_row=sh.max_row, values_only=True):
            if row[4] is not None:
                if row[27] in RISK:
                    RISK[row[27]] += 1

        print(f'{subject_name} - {RISK}')
        wb_errors = openpyxl.Workbook()
        sh_errors = wb_errors.worksheets[0]
        sh_errors.append(('субъект', "чрезвычайно высокий риск", "высокий риск", "значительный риск", "средний риск", "умеренный риск", "низкий риск"))
        sh_errors.append((subject_name, RISK["чрезвычайно высокий риск"], RISK["высокий риск"], RISK["значительный риск"], RISK["средний риск"], RISK["умеренный риск"], RISK["низкий риск"]))
        wb_errors.save(f'{dir}/ошибки риск {subject_name}.xlsx')
        return RISK

    def coint_deyatel(self, dir, start_from_row):

        file_name = str(dir.split('/')[-1])

        count = 0

        if file_name == '.DS_Store':
            return 0
        else:
            file = os.listdir(dir)[0]
            wb = openpyxl.load_workbook(f'{dir}/{file}')
            sh = wb.worksheets[0]

            for n, ul in enumerate(sh['B']):
                find_match = False
                if n < start_from_row:
                    continue
                ul_name = ul.value
                if ul_name != None:
                    # print(f'{n} {ul_name}')
                    level_risk = sh.cell(row=n+1, column=28).value
                    if level_risk == 'чрезвычайно высокий риск':

                        kind_deyat = sh.cell(row=n+1, column=6).value

                        for chvr in CHVR:
                            find_match = False

                            for word in chvr['word']:
                                if chvr['name'] == 'kosmos':
                                    if re.findall(word, str(kind_deyat).upper()):
                                        chvr['count'] += 1
                                        chvr['knm'].append(sh.cell(row=n + 1, column=32).value)
                                        find_match = True
                                        break
                                elif chvr['name'] == 'school_lager':

                                    variants = [chvr['word'] for chvr in CHVR if chvr['name'] == 'shkola'][0]

                                    for v in variants:
                                        if re.findall(v, str(ul_name).upper()):
                                            if re.findall(word, str(kind_deyat).upper()) or re.findall(word, str(ul_name).upper()):
                                                chvr['count'] += 1
                                                chvr['knm'].append(sh.cell(row=n + 1, column=32).value)
                                                find_match = True
                                                break


                                else:
                                    if re.findall(word, str(kind_deyat).upper()) or re.findall(word, str(ul_name).upper()):
                                        chvr['count'] += 1
                                        chvr['knm'].append(sh.cell(row=n+1, column=32).value)
                                        find_match = True
                                        break

                            if find_match is True:
                                break
            for c in CHVR:
                print(f"{c['name']} {c['count']}")

            return CHVR




    def merge_files(self, files_list):
        # print(abs_path_file)

        abs_path_file_0 = f'{self.dir}/{files_list[0]}'
        abs_path_file_1 = f'{self.dir}/{files_list[1]}'

        wb_0 = openpyxl.load_workbook(abs_path_file_0)

        wb_1 = openpyxl.load_workbook(abs_path_file_1)

        sh_0 = wb_0.worksheets[0]
        sh_1 = wb_1.worksheets[0]

        sh_0_max_row = sh_0.max_row
        print(f"{sh_0_max_row}")

        for row_1 in sh_1.iter_rows(min_row=self.start_from_row, max_row=sh_1.max_row, values_only=True):
            if row_1.count(None) > 40:
                continue
            else:
                sh_0.append(row_1)
        # end_empty_row = self.start_from_row
        # for n, row_0 in enumerate(sh_0.iter_rows(min_row=self.start_from_row, max_row=sh_0.max_row, values_only=True)):
        #
        #     if row_0.count(None) > 40:
        #         end_empty_row += 1
        #
        # amount = end_empty_row - self.start_from_row
        # if amount > 0:
        #     sh_0.delete_rows(idx=self.start_from_row, amount=amount)


        wb_0.save(abs_path_file_0)
        wb_1.close()
        os.remove(abs_path_file_1)

        # print(f'{len(column_B)=} {len(column_D)=}')



    def open_lib(self):
        wb_lib = openpyxl.load_workbook(self.path_file_lib)
        sh_lib = wb_lib.worksheets[0]
        column_plans = sh_lib['A']
        column_regions = sh_lib['C']
        for n, cl in enumerate(column_plans):
            r = self.list_files(number=cl.value)
            if r is not None:
                region_name = column_regions[n].value
                c = self.create_new_dir(region_name)
                shutil.copy(r, c)



    def list_files(self, number):

        file_path = f'{self.path_files}/{number}_полный_план.xlsx'
        if os.path.exists(file_path):
            return file_path
        else:
            return None

    def create_new_dir(self, dir_name):
        dir_name_path = f'/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/{dir_name}'
        if not os.path.exists(dir_name_path):
            os.mkdir(dir_name_path)
        return dir_name_path






if __name__ == '__main__':
    Exel_work().sort_from_lib()
