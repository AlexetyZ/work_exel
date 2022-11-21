import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from pathlib import Path


class Compair:

    def __init__(self, main_file_path: str,
                 compared_file_path: str,
                 startswith_row_main: int,
                 startswith_row_compared: int,
                 operative_column_main: str = 'I',
                 alter_operative_column_main: str = 'AH',
                 operative_column_compare: str = 'Y',
                 alter_operative_column_compare: str = 'AH',
                 info_column_compare: str = 'B',
                 record_column_main: str = 'AH',
                 color_for_info_column_compare: str = "ffff00"
                 ):

        """
        В кратце, и по сути! Берет главный файл, берет сравниваемый файл, и при найденном совпадении в указанных столбцах, берет данные
        из сравниваемого файла и вставляет в главный файл напротив искомого результата с соединением через перенос строки

        получается, что главный файл теоретически меньше сравниваемоого, а напротив искомых ячеек суммируются строки
        при найденных в сравниваемом файле ячейках

        :param main_file_path: путь до файла, из которого будем брать данные для сравнения (далее - главный файл)
        :param compared_file_path: путь до файла, значения из которого будем сравнивать (далее - сравниваемый файл)
        :param startswith_row_main: с какой строки заканчивается шапка таблицы и начинаются данные в главном файле
        :param startswith_row_compared: с какой строки заканчивается шапка таблицы и начинаются данные в сравниваемом файле
        :param operative_column_main: буква столбца главного файла, из которого будем брать данные для сравнения
        :param alter_operative_column_main: альтернативная (при отсутствии значения в основной) буква столбца главного файла, из которого будем брать данные для сравнения
        :param operative_column_compare: буква столбца сравниваемого файла, значения которого будем сравнивать
        :param alter_operative_column_compare: альтернативная (при отсутствии значения в основной) буква столбца сравниваемого файла, значения которого будем сравнивать
        :param info_column_compare: буква столбца сравниваемого файла, значения которого будут браться при найденном совпадении
        :param record_column_main: буква столбца главного файла, куда будем записывать найденные данные из сравниваемой таблицы
        :param color_for_info_column_compare: цвет ячейки столбца сравниваемого файла в случае совпадения


        """
        self.main_file_path = main_file_path
        self.main_wb = openpyxl.load_workbook(main_file_path)
        self.main_sh = self.main_wb.worksheets[0]
        self.operative_column_main = operative_column_main
        self.alter_operative_column_main = alter_operative_column_main
        self.main_column = self.main_sh[operative_column_main]
        self.startswith_row_main = startswith_row_main - 1   # 3
        self.record_column_main = record_column_main

        self.compared_file_path = compared_file_path
        self.compared_file_name = Path(self.compared_file_path).stem
        self.compared_wb = openpyxl.load_workbook(compared_file_path)
        self.compared_sh = self.compared_wb.worksheets[0]
        self.compared_column = self.compared_sh[operative_column_compare]
        self.startswith_row_compared = startswith_row_compared - 1   # 1
        self.info_column_compare = info_column_compare
        self.alter_operative_column_compare = alter_operative_column_compare
        self.color_for_info_column_compare = color_for_info_column_compare


        self.compair()

    def compair(self):
        try:
            for main_row_number, main_cell in enumerate(self.main_column):
                objects_addresses = ''
                if main_row_number < self.startswith_row_main:
                    continue
                if main_cell.value is None:
                    main_ogrn = str(self.main_sh[f'{self.alter_operative_column_main}{main_row_number+1}'].value).strip()
                else:
                    main_ogrn = str(main_cell.value).strip()
                print(f'{main_row_number} {main_ogrn}')

                for compared_row_number, compared_cell in enumerate(self.compared_column):
                    if compared_row_number < self.startswith_row_compared:
                        continue
                    if compared_cell.value is None:

                        compared_ogrn = str(self.compared_sh[f'{self.alter_operative_column_compare}{compared_row_number+1}'].value).strip()
                    else:
                        compared_ogrn = str(compared_cell.value).strip()





                    if main_ogrn == compared_ogrn:
                        print(f'совпадение в строке {compared_row_number + 1}')
                        info_compare_cell = self.compared_sh[f'{self.info_column_compare}{compared_row_number + 1}']
                        info_compare_cell.fill = PatternFill('solid', fgColor=self.color_for_info_column_compare)
                        objects_addresses = objects_addresses + str(info_compare_cell.value).strip() + f" (файл {self.compared_file_name}, строка{compared_row_number + 1})" + "\n"

                        if self.main_sh[f'{self.record_column_main}{main_row_number+1}'].value is not None:
                            self.main_sh[f'{self.record_column_main}{main_row_number+1}'].value = self.main_sh[f'{self.record_column_main}{main_row_number+1}'].value + "\n" + objects_addresses
                        else:
                            self.main_sh[f'{self.record_column_main}{main_row_number + 1}'].value = objects_addresses
            self.main_wb.save(self.main_file_path)
            self.compared_wb.save(self.compared_file_path)

        except KeyboardInterrupt:
            print('остановка программы и сохранение последних изменений...')
            self.main_wb.save(self.main_file_path)
            self.compared_wb.save(self.compared_file_path)
        except Exception as ex:
            print(f'Некая ошибка, которую не удалось опознать:{ex}')
            self.main_wb.save(self.main_file_path)
            self.compared_wb.save(self.compared_file_path)


if __name__ == '__main__':

    main_file_path = '/Volumes/KINGSTON/Выверка ЕРВК/РПТН_04.xlsx'
    compared_file_path = '/Volumes/KINGSTON/Выверка ЕРВК/Федеральная служба по надзору в сфере защиты прав потребителей и благополучия человека Часть 1.xlsx'
    Compair(
        main_file_path=main_file_path,
        compared_file_path=compared_file_path,
        startswith_row_main=9790,
        startswith_row_compared=2,
        operative_column_main='X',
        alter_operative_column_main='AH',
        operative_column_compare='H',
        alter_operative_column_compare='D',
        info_column_compare='J',
        record_column_main='AZ'
    )
# последняя обработанная строчка - 9790

