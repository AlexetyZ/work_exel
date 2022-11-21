import re

from pxl import Exel_work
import os
import openpyxl

main_dir = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям' # папка, где будут готовые отсортированнные файлы
path_file_lib = '/Users/aleksejzajcev/Desktop/Список_планов_КНМ.xlsx' # файл, содержащий список выгруженных файлов
path_files: str = '/Users/aleksejzajcev/Desktop/выгруженные планы на 2023' # папка, содержащая файлы выгруженных планов
start_from_row = 19 #  номер строчки с которой заканчиваются заголовки таблицы и начинается содержимое таблицы в файлах планов

def find_errors_risk():

    wb_main = openpyxl.Workbook()
    sh_main = wb_main.worksheets[0]
    sh_main.append(('субъект', "чрезвычайно высокий риск", "высокий риск", "значительный риск", "средний риск", "умеренный риск", "низкий риск"))
    for dir in os.listdir(main_dir):
        risk = Exel_work().count_risk(f'{main_dir}/{dir}', start_from_row=start_from_row)
        if risk == 0:
            continue
        sh_main.append((dir, risk["чрезвычайно высокий риск"], risk["высокий риск"], risk["значительный риск"],
                          risk["средний риск"], risk["умеренный риск"], risk["низкий риск"]))

    wb_main.save(f'{main_dir}/общая выгрузка по рискам.xlsx')


def delete_errors_in_all_dirs():

    for dir in os.listdir(main_dir):
        if os.path.exists(f'{main_dir}/{dir}/ошибки риск {dir}.xlsx'):
            os.remove(f'{main_dir}/{dir}/ошибки риск {dir}.xlsx')


def sort_plan_by_dirs():
    """
    Сортирует файлы планов проверок по папкам. Необходима папка на работем столе "выгрузка файлов по областям", а так же не забудь в объявленных переменных изменить полный путь
    path_files и  path_file_lib соответствено
    :return:
    """
    Exel_work().sort_from_lib(path_files, path_file_lib)


def merge_plans_in_dirs():
    """
    объединяет однотипные файлы в директории папок, подчищая последнюю строку, по помни, нужно иметь показательную
    ячейку, пустота в которой будет означать, что вся строка пустая.

    :return:
    """

    for dir in os.listdir(main_dir):
        Exel_work().merge_files_in_directory(f'{main_dir}/{dir}', start_from_row=start_from_row)


def merge_plans_in_dir(dir):
    """
    объединяет однотипные файлы в указанной папке, подчищая последнюю строку, по помни, нужно иметь показательную
    ячейку, пустота в которой будет означать, что вся строка пустая.

    :return:
    """
    Exel_work().merge_files_in_directory(dir, start_from_row=start_from_row)


def delete_last_empty_rows_in_file(file_path):
    Exel_work().delete_last_empty_rows(file_path, start_from_row=start_from_row)





def count_subjects_and_objects():
    """
    Считает количество субъектов и обектов в плане проверок по количеству заполненных ячеек ИНН и ОРГН. Их число говорит
    о количестве объектов, а множество "set()" говорит о количестве субъектов.
    Проходится в указанной директории по папкам, именованных территориальными управлениями и выбирает файлы, в названии
    которых есть "+". Необходимо убедиться, чтобы первый столбец был заполнен, так как по нему идет опредленение
    последней строчки в файле.
    По окончании выполнения скрипта, создается файл exel "не включенные по субъектам.xlsx" со столбцами
    "ТУ", "Не включенных объектов", "Не включенных субъектов", где под ТУ будет название папки в директории.
    :return:
    """
    dir_list_path = '/Volumes/KINGSTON/Субъекты в ЕРВК'
    dirs_list = os.listdir(dir_list_path)
    main_wb = openpyxl.Workbook()
    main_sh = main_wb.worksheets[0]
    main_sh.append(("ТУ", "Не включенных объектов", "Не включенных субъектов"))


    for dir in dirs_list:
        file_list = os.listdir(f'{dir_list_path}/{dir}')
        no_files = True
        for file in file_list:


            if re.search("\+", file) and not re.search("\._", file):
                no_files = False
                subject = []
                wb = openpyxl.load_workbook(f'{dir_list_path}/{dir}/{file}')
                sh = wb.worksheets[0]
                for row in sh.iter_rows(min_row=1, values_only=True):
                    if row[0] is None:
                        break
                    try:
                        srt_int = int(row[23])
                        subject.append(row[23])

                    except:
                        try:
                            srt_int = int(row[33])
                            subject.append(row[33])

                        except:
                            continue


                print(f'{dir}, {len(subject)}, {len(set(subject))}')
                main_sh.append((dir, len(subject), len(set(subject))))
                print('')
                break


        if no_files == True:

            main_sh.append((dir, "без нарушений", "без нарушений"))
            print(dir, "без нарушений", "без нарушений")
            print('')

    main_wb.save('/Users/aleksejzajcev/Desktop/не включенные по субъектам.xlsx')


def find_disappear_plans_numbers():
    """
    сравнивает значения в списке планов проверок и директории загруженных проверок и выявляется недостающие проверки
    согласно списку планов.
    :return:
    Список номеров планов проверок, отсутствующих в директории выгруженных планов
    Если в директории планов имеются все файлы, то возвращается пустой список
    """
    disappear = []
    plans = []

    plans_list_dir = os.listdir(path_files)
    for plan in plans_list_dir:
        plans.append(re.split('_', plan)[0])
    # print(plans)

    wb = openpyxl.load_workbook(path_file_lib)
    sh = wb.worksheets[0]
    for number in sh['A']:
        if number.value not in plans:
            disappear.append(number.value)

    return disappear


def main():

    print(find_disappear_plans_numbers())
    # sort_plan_by_dirs()
    # merge_plans_in_dir('/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/РПН Республике Дагестан')
    # merge_plans_in_dirs()
    # Exel_work().delete_last_empty_rows('/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/РПН Республике Дагестан/2023049877_полный_план.xlsx', start_from_row)
    pass


if __name__ == '__main__':
    main()


# dire = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/Межрегиональное РПН Республике Крым и городу федерального значения Севастополю'
# Exel_work().coint_knm(dire, 18)
# Exel_work().coint_deyatel(dire, 18)

# dir_TU = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/Межрегиональное РПН Республике Крым и городу федерального значения Севастополю'
#
# Exel_work().count_risk(dir_TU, 18)


# dir = '/Users/aleksejzajcev/Desktop/выгрузка файлов по областям/РПН железнодорожному транспорту'
# Exel_work().sort_from_lib()